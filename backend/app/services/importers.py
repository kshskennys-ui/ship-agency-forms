import json
from datetime import date, datetime, timedelta
from pathlib import Path


def excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (date(1899, 12, 30) + timedelta(days=int(value)))
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            pass
    return None


def clean_code(value):
    if value is None:
        return None
    text = str(value).strip()
    return text.split("-", 1)[1] if "-" in text else text


def _header_map(headers):
    aliases = {
        "name": ("姓名", "船员姓名", "family name"),
        "gender": ("性别", "m/f"),
        "nationality": ("国籍", "国藉", "nationality"),
        "rank": ("职务", "职别", "rank"),
        "birth_date": ("出生日期", "date of birth"),
        "document_no": ("证件号", "证书号", "证件号码", "passport"),
    }
    mapped = {}
    for key, options in aliases.items():
        for index, header in enumerate(headers):
            text = str(header or "").lower()
            if any(option.lower() in text for option in options):
                mapped[key] = index
                break
    return mapped


def parse_crew_file(path: str):
    extension = Path(path).suffix.lower()
    if extension == ".xls":
        import xlrd
        book = xlrd.open_workbook(path, on_demand=True)
        try:
            sheet = book.sheet_by_index(0)
            rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        finally:
            book.release_resources()
    else:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return [], {"source_type": extension[1:]}
    header_index = next((i for i, row in enumerate(rows[:8]) if any("姓名" in str(x or "") for x in row)), 0)
    headers = rows[header_index]
    mapping = _header_map(headers)
    if "name" not in mapping:
        raise ValueError("未找到船员姓名字段")
    members = []
    for row in rows[header_index + 1:]:
        name = row[mapping["name"]] if mapping["name"] < len(row) else None
        if name in (None, ""):
            continue
        members.append({
            "name": str(name).strip(),
            "gender": clean_code(row[mapping["gender"]]) if "gender" in mapping else None,
            "nationality": clean_code(row[mapping["nationality"]]) if "nationality" in mapping else None,
            "rank": clean_code(row[mapping["rank"]]) if "rank" in mapping else None,
            "birth_date": excel_date(row[mapping["birth_date"]]) if "birth_date" in mapping else None,
            "document_no": str(row[mapping["document_no"]]).strip() if "document_no" in mapping and row[mapping["document_no"]] not in (None, "") else None,
            "extra": {"raw_row": [str(x) for x in row]},
        })
    return members, {"source_type": extension[1:], "header_row": header_index}
